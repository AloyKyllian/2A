import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import os

path_cavalier_mercredi = "liste_cavalier_mercredi.txt"
path_cavalier_samedi = "liste_cavalier_samedi.txt"


def colorier_ancient_chevaux(ancient_cheval_eleve):
    if len(ancient_cheval_eleve) >= 3:
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[2][1], {'bg': 'yellow'})
    if len(ancient_cheval_eleve) >= 2:
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[1][1], {'bg': 'orange'})
    if len(ancient_cheval_eleve) >= 1:
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[0][1], {'bg': 'red'})


def heure_precedant():
    liste_heure = list(plannig1.liste_heure)
    for i in range(1, len(liste_heure)):
        if liste_heure[i] == cellule.heu:
            heure.set_heure(liste_heure[i-1])


def heure_suivant():
    liste_heure = list(plannig1.liste_heure)
    for i in range(0, len(liste_heure)-1):
        if liste_heure[i] == cellule.heu:
            heure.set_heure(liste_heure[i+1])
            return 0


def inserer_listbox(i):
    if len(plannig1.liste) != 0:
        present = any((heure.h, i) == (tup[0], tup[2])
                      for tup in plannig1.liste)
        if present:
            eleve_listbox.insert(tk.END, i)
            eleve_listbox.itemconfig(tk.END, {'bg': 'red'})
        else:
            eleve_listbox.insert(tk.END, i)
    else:
        eleve_listbox.insert(tk.END, i)


def ajouteleve():
    vider_listebox(eleve_listbox)
    for i in plannig1.liste_personne[heure.h]:
        inserer_listbox(i)


def colorier():
    # print(plannig1.cheval)
    for i in range(0, len(plannig1.cheval)):
        cheval_listbox.itemconfig(
            i, {'bg': 'white'})
    setcheval = set()
    for i in plannig1.liste_personne[heure.h]:
        ancient = plannig1.ancient_cheval_de(i, heure.h)
        for y in ancient:
            setcheval.add(y[1])
    for i in setcheval:
        cheval_listbox.itemconfig(
            i, {'bg': 'green'})


def vider_listebox(listebox):
    if listebox.size() > 0:
        listebox.delete(0, listebox.size())


def ajoutcheval():
    cheval_listbox.delete(0, END)
    for i in plannig1.liste_chevaux:
        cheval_listbox.insert(tk.END, i)
    if heure.h:
        colorier()


def ajouter():
    possible = True
    if heure.h != "heure" and cellule.cheval != "cheval" and cellule.eleve != "cavalier":
        for i in plannig1.liste:
            if (i[0], i[1]) == (heure.h, cellule.cheval):
                possible = False

        if possible == True:
            if plannig1.heure_travailler(cellule.cheval) < 4:
                i = plannig1.liste_chevaux[cellule.ind_cheval]
                plannig1.liste_chevaux.remove(i)
                plannig1.liste_chevaux.insert(
                    cellule.ind_cheval, (i[0]+1, i[1]))
                ajoutcheval()
                if cellule.ind_eleve[0] != None:
                    eleve_listbox.itemconfig(
                        cellule.ind_eleve[0], {'bg': 'red'})
                Nb = 0
                for i in range(cheval_listbox.size()):
                    if cellule.cheval == cheval_listbox.get(i)[1]:
                        cellule.ind_cheval = Nb
                    Nb = Nb+1
                plannig1.ajout(heure.h, cellule.cheval, cellule.eleve)
                visu_fichier.delete('1.0', END)
                visu_fichier.insert(END, plannig1.fichier(journee.j))
                label_enregistrer.config(fg="#f0f0f0")
                inserer_liste_de_travaille()
                plannig1.append_historique(
                    "ajout", (heure.h, cellule.cheval, cellule.eleve))
            else:
                msgbox = tk.messagebox.showerror(
                    title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja 4 heure dans la journée")
        else:
            msgbox = tk.messagebox.showerror(
                title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja durant cette heure")
    else:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="vous n'avez pas selectionné toutes les informations necessaire à l'ajout")


def changer_heure():
    vider_listebox(eleve_listbox)
    ajouteleve()
    colorier()
    varheure.set(f"{heure.h}")
    cellule.set_heure(heure.h)
    varajout.set(cellule.getCellule())


def supprimer():
    if (heure.h, cellule.cheval, cellule.eleve) in plannig1.liste:
        plannig1.liste.remove((heure.h, cellule.cheval, cellule.eleve))
        i = plannig1.liste_chevaux[cellule.ind_cheval]
        plannig1.liste_chevaux.remove(i)
        plannig1.liste_chevaux.insert(cellule.ind_cheval, (i[0]-1, i[1]))
        ajoutcheval()
        if cellule.ind_eleve[0] != None:
            eleve_listbox.itemconfig(cellule.ind_eleve[0], {'bg': 'white'})
        Nb = 0
        for i in range(cheval_listbox.size()):
            if cellule.cheval == cheval_listbox.get(i)[1]:
                cellule.ind_cheval = Nb
            Nb = Nb+1
        visu_fichier.delete('1.0', END)
        visu_fichier.insert(END, plannig1.fichier(journee.j))
        label_enregistrer.config(fg="#f0f0f0")
        inserer_liste_de_travaille()
        plannig1.append_historique(
            "suppression", (heure.h, cellule.cheval, cellule.eleve))


def inserer_liste_de_travaille():
    vider_listebox(heure_listebox)
    liste = plannig1.heure_cavalier_de(cellule.cheval)
    for i in liste:
        heure_listebox.insert(tk.END, i)


def cmp_dates(d):
    j, m, a = d[0].split("-")
    return (int(a), int(m), int(j))


def recup_donne():
    plannig1.lire_fichiers()
    ajout_des_commande_lena()
    msgbox = tk.messagebox.showinfo(
        title="selection de fichier", message="veillez selectionner le fichier que vous voulait remplir")
    varjour.set(journee.j)
    chemin = tk.Tk()
    chemin.withdraw()                 # pour ne pas afficher la fenêtre Tk
    name = askopenfilename()

    plannig1.recupperation_excel("", name)
    folder = name.split('/')
    name = folder[-1]
    folder = folder[:-1]
    path = '/'.join(folder)
    liste = []
    for files in os.listdir(path):
        if files != name and ".xlsx" in files and '$' not in files:
            if journee.j == "Mercredi" and "mercredi" in files:
                liste.append((files[15:25], files))
            elif journee.j == "Samedi" and "samedi" in files:
                liste.append((files[13:23], files))
    liste = sorted(liste, key=cmp_dates, reverse=True)
    if len(liste) >= 1:
        plannig1.recupperation_excel("ancien", path + '/' + liste[0][1])
    if len(liste) >= 2:
        plannig1.recupperation_excel("ancien2", path + '/' + liste[1][1])
    if len(liste) >= 3:
        plannig1.recupperation_excel("ancien3", path + '/' + liste[2][1])
    msgbox = tk.messagebox.showinfo(
        title="creation de fichier", message="tout les fichiers ont été recuperé")


def recup_donne_mercredi():
    journee.setmercredi()
    recup_donne()


def recup_donne_samedi():
    journee.setsamedi()
    recup_donne()


class Jour():

    def __init__(self):
        self.j = ""

    def setmercredi(self):
        self.j = "Mercredi"

    def setsamedi(self):
        self.j = "Samedi"


class Heure():

    def __init__(self):
        self.h = ""

    def set_heure(self, h):
        self.h = h
        changer_heure()

    def h1(self):
        if journee.j == "Mercredi":
            self.h = "10H LENA"
        else:
            self.h = "9h30 L"
        changer_heure()

    def h2(self):
        if journee.j == "Mercredi":
            self.h = "13H30 LENA"
        else:
            self.h = "9H30 K"
        changer_heure()

    def h3(self):
        if journee.j == "Mercredi":
            self.h = "14H LENA"
        else:
            self.h = "10H30 L"
        changer_heure()

    def h4(self):
        if journee.j == "Mercredi":
            self.h = "15H LENA"
        else:
            self.h = "10H30 K"
        changer_heure()

    def h5(self):
        if journee.j == "Mercredi":
            self.h = "16H LENA "
        else:
            self.h = "11H30 L"
        changer_heure()

    def h6(self):
        if journee.j == "Mercredi":
            self.h = "17H LENA"
        else:
            self.h = "11H30 K"
        changer_heure()

    def h7(self):
        if journee.j == "Mercredi":
            self.h = "18H LENA "
        else:
            self.h = "12H00 L"
        changer_heure()

    def h8(self):
        if journee.j == "Mercredi":
            self.h = "19H LENA "
        else:
            self.h = "13H00 K"
        changer_heure()

    def h9(self):
        if journee.j == "Mercredi":
            self.h = "15H KARINE"
        else:
            self.h = "14H00 L"
        changer_heure()

    def h10(self):
        if journee.j == "Mercredi":
            self.h = "17H KARINE"
        else:
            self.h = "14H00 K"
        changer_heure()

    def h11(self):
        if journee.j == "Mercredi":
            self.h = "18H KARINE"
        else:
            self.h = "15H00 L"
        changer_heure()

    def h12(self):
        if journee.j == "Mercredi":
            self.h = "19"
        else:
            self.h = "16H00 K"
        changer_heure()

    def h13(self):
        self.h = "16H00 L"
        changer_heure()

    def h14(self):
        self.h = "17H00 K"
        changer_heure()

    def h15(self):
        self.h = "17H00 L"
        changer_heure()

    def h16(self):
        self.h = "18H00 K "
        changer_heure()


class Cellule():

    def __init__(self):
        self.heu = "heure"
        self.eleve = "cavalier"
        self.cheval = "cheval"
        self.ind_eleve = 0
        self.ind_che = 0

    def set_heure(self, heure):
        self.heu = heure

    def set_ind_cava(self, ind):
        self.ind_eleve = ind

    def set_ind_che(self, ind):
        self.ind_cheval = ind

    def set_cavalier(self, cavalier):
        self.eleve = cavalier

    def set_cheval(self, cheval):
        self.cheval = cheval
        self.ind_cheval = plannig1.index_cheval(cellule.cheval)

    def getCellule(self):
        return ((self.heu, self.cheval, self.eleve))


class plannig():

    def __init__(self):

        # self.liste_heure = ["10H LENA",	"13H30 LENA", "14H LENA", "15H KARINE", "15H LENA",
        #                              "16H LENA", "17H LENA",	"17H KARINE", "18H LENA", "18H KARINE",	"19H LENA"]

        # self.liste_heure_samedi =
        self.historique = []  # (type, (heure,che,cava))

        self.name_ancien_fichier = ""
        self.name_fichier = ""
        self.liste_heure = {}
        self.cheval = {}

        self.liste_personne = {}
        self.liste_personne_sam = {}

        self.liste_chevaux = []
        self.liste = []
        self.ancien_liste = []
        self.ancien_liste2 = []
        self.ancien_liste3 = []

    def append_historique(self, type, donne):
        if len(self.historique) > 100:
            self.historique = self.historique[10:]
        self.historique.append((type, donne))
        historique.delete('1.0', END)
        for i in self.historique:
            historique.insert("1.0", f"{i}\r\n")

    def lire_fichiers(self):

        if journee.j == "Mercredi":
            fichier_mer = open(path_cavalier_mercredi, "r")
            lignes = fichier_mer.read()
            lignes = lignes.split("\n")
            liste = []
            for ligne in lignes:
                if "H LENA" in ligne or "H30 LENA" in ligne or "H KARINE" in ligne:
                    self.liste_personne[ligne] = liste[:]
                    liste.clear()
                else:
                    liste.append(ligne)
            fichier_mer.close()

        elif journee.j == "Samedi":
            fichier_sam = open(path_cavalier_samedi, "r")
            lignes = fichier_sam.read()
            lignes = lignes.split("\n")
            liste = []
            for ligne in lignes:
                if "H00" in ligne or "30" in ligne:
                    self.liste_personne[ligne] = liste[:]
                    liste.clear()
                else:
                    liste.append(ligne)
            fichier_sam.close()

    def fichier(self, jour):
        heure = 0
        a = 0
        txt = ""
        txt = (f"\t\t\t\t\t\tPlanning {jour}")
        for i in self.liste:
            if heure != i[0]:
                heure = i[0]
                txt = txt + (f"\r\n\r\n{i[0]} :\t ")
            txt = txt + (f"{i[2]} avec {i[1]} | ")
            if a == 1:
                a = -1
                txt += "\r\n\t    "
            a += 1
        return txt

    def ajout(self, heure, cheval, personne):
        self.liste.append((heure, cheval, personne))
        self.liste.sort()

    def index_cheval(self, cheval):
        return self.cheval[cheval]-4

    def ancient_cheval_de(self, personne, heure):
        cavalier = []
        for i in self.ancien_liste:
            if (i[2], i[0]) == (personne, heure):
                cavalier.append((i[1], self.cheval[i[1]]-4))
        for i in self.ancien_liste2:
            if (i[2], i[0]) == (personne, heure):
                cavalier.append((i[1], self.cheval[i[1]]-4))
        for i in self.ancien_liste3:
            if (i[2], i[0]) == (personne, heure):
                cavalier.append((i[1], self.cheval[i[1]]-4))
        return cavalier

    def ancient_cavalier_de(self, cheval):
        liste = []
        for i in self.ancien_liste:
            if i[1] == cheval:
                liste.append((i[2], i[0]))
        return liste

    def cavalier_de(self, cheval):
        for i in self.liste:
            if i[1] == cheval:
                yield i[2]

    def heure_de(self, cheval):
        for i in self.liste:
            if i[1] == cheval:
                yield i[0]

    def heure_cavalier_de(self, cheval):
        for i in self.liste:
            if i[1] == cheval:
                yield (i[0], i[2])

    def heure_travailler(self, cheval):
        nbr = 0
        for i in self.liste:
            if i[1] == cheval:
                nbr = nbr+1
        return nbr

    def ecrire_fichier(self):
        workbook = load_workbook(self.name_fichier)
        sheet = workbook.active
        for ligne in range(4, len(self.cheval)+4):
            for colonne in range(2, len(self.liste_heure)+2):
                sheet.cell(ligne, colonne).value = None
        for i in self.liste:
            if sheet.cell(self.cheval[i[1]], self.liste_heure[i[0]]).value == None:
                sheet.cell(self.cheval[i[1]],
                           self.liste_heure[i[0]]).value = i[2]
        err = workbook.save(self.name_fichier)
        if err == None:
            label_enregistrer.config(fg="#000000")

    def recupperation_excel(self, listeself, name):
        if listeself == "":
            self.name_fichier = name  # lance la fenêtre
        else:
            self.name_ancien_fichier = name
        workbook = load_workbook(name)
        sheet = workbook.active
        liste_che = {}
        liste = []
        liste_heure = {}

        for j in range(1, 18):
            for i in range(3, 70):
                if sheet.cell(row=i, column=j).value != None:
                    if i == 3:
                        if listeself == "":
                            self.liste_heure[sheet.cell(
                                row=i, column=j).value] = j
                        liste_heure[j] = sheet.cell(row=i, column=j).value
                    elif j == 1:
                        if listeself == "":
                            self.cheval[sheet.cell(row=i, column=j).value] = i
                        liste_che[i] = sheet.cell(row=i, column=j).value
                    elif j > 1 and sheet.cell(row=i, column=j).value != "MERCREDI" and sheet.cell(row=i, column=j).value != "SAMEDI":
                        liste.append(
                            (liste_heure[j], liste_che[i], sheet.cell(row=i, column=j).value))
        if listeself == "ancien":
            self.ancien_liste = liste
        elif listeself == "ancien2":
            self.ancien_liste2 = liste
        elif listeself == "ancien3":
            self.ancien_liste3 = liste
        elif listeself == "":
            self.liste = liste
            plannig1.liste_chevaux.clear()
            for i in sheet["A"]:
                if i.value != None:
                    self.liste_chevaux.append(
                        (self.heure_travailler(i.value), i.value))
            ajoutcheval()
            visu_fichier.delete('1.0', END)
            visu_fichier.insert(END, plannig1.fichier(journee.j))

    def recup_donne_samedi(self):
        self.lire_fichiers()
        journee.setsamedi()
        ajout_des_commande_lena()
        msgbox = tk.messagebox.showinfo(
            title="selection de fichier", message="veillez selectionner le fichier que vous voulait remplir")
        self.recupperation_excel("")
        msgbox = tk.messagebox.showinfo(
            title="creation de fichier", message="veillez selectionner l'ancient fichier de la semaine dernière")
        self.recupperation_excel("ancien")
        self.recupperation_excel("ancien2")
        self.recupperation_excel("ancien3")

    def ancien_cavalier(self, cheval):
        liste = []
        for i in self.liste:
            liste.append(i[1])
        return liste


journee = Jour()
plannig1 = plannig()
# Création de l'interface utilisateur
window = tk.Tk()
window.title("Planning")
window.attributes('-fullscreen', True)
window.bind('<Escape>', lambda e: window.destroy())
frame = tk.Frame(master=window, width=300, height=100)

frame.pack()

cellule = Cellule()

heure = Heure()

varheure = StringVar()
varjour = StringVar()
varcavalier = StringVar()
varcheval = StringVar()
varajout = StringVar()
varheure_cheval = StringVar()
varcavalier1 = StringVar()
varcavalier2 = StringVar()

label_jour = tk.Label(window, textvariable=varjour)
label_jour.place(x=20, y=20)

label_heure = tk.Label(window, textvariable=varheure)
label_heure.place(x=70, y=20)

boutton_avancer_heure = tk.Button(
    window, width=2, text="<", command=heure_precedant)
boutton_avancer_heure.place(x=150, y=20)

boutton_reculer_heure = tk.Button(
    window, width=2, text=">", command=heure_suivant)
boutton_reculer_heure.place(x=180, y=20)


label_cavalier = tk.Label(window, text="INFOS cavalier")
label_cavalier.place(x=450, y=70)

label_cavalier2 = tk.Label(
    window, text="la semaine derniere il/elle a monté : ")
label_cavalier2.place(x=450, y=100)

label_cavalier3 = tk.Label(
    window, textvariable=varcavalier)
label_cavalier3.place(x=650, y=100)

label_cavalier6 = tk.Label(
    window, text="il y a 2 semaines il/elle a monté : ")
label_cavalier6.place(x=450, y=120)

label_cavalier4 = tk.Label(
    window, textvariable=varcavalier1)
label_cavalier4.place(x=650, y=120)

label_cavalier7 = tk.Label(
    window, text="il y a 3 semaines il/elle a monté : ")
label_cavalier7.place(x=450, y=140)

label_cavalier5 = tk.Label(
    window, textvariable=varcavalier2)
label_cavalier5.place(x=650, y=140)

varcavalier.set("cheval")
varcavalier1.set("cheval1")
varcavalier2.set("cheval2")

eleve_listbox = tk.Listbox(window, yscrollcommand=True)
eleve_listbox.place(x=20, y=50)


def items_selected(event):
    # get all selected indices
    selected_indices = eleve_listbox.curselection()
    cellule.set_ind_cava(selected_indices)
    cellule.set_cavalier(eleve_listbox.get(selected_indices))
    ancient_cheval = plannig1.ancient_cheval_de(
        cellule.eleve, heure.h)
    if len(ancient_cheval) >= 1:
        varcavalier.set(ancient_cheval[0][0])
    else:
        varcavalier.set("cheval")
    if len(ancient_cheval) >= 2:
        varcavalier1.set(ancient_cheval[1][0])
    else:
        varcavalier1.set("cheval1")
    if len(ancient_cheval) >= 3:
        varcavalier2.set(ancient_cheval[2][0])
    else:
        varcavalier2.set("cheval2")
    colorier()
    colorier_ancient_chevaux(ancient_cheval)
    for tup in plannig1.liste:
        if (heure.h, cellule.eleve) == (tup[0], tup[2]):
            cavalier = []
            cellule.set_cheval(tup[1])
            ancient_cavalier = plannig1.ancient_cavalier_de(cellule.cheval)
            for i in ancient_cavalier:
                cavalier.append(f"{i[0]} a {i[1]}")
            varheure_cheval.set(f"heure de travaille de : {cellule.cheval}")
            varcheval.set(cavalier)
            inserer_liste_de_travaille()
    varajout.set(cellule.getCellule())


eleve_listbox.bind('<ButtonRelease-1>', items_selected)

label_cheval = tk.Label(window, text="INFOS cheval")
label_cheval.place(x=450, y=300)

label_cheval2 = tk.Label(
    window, text="la semaine derniere il/elle a ete monte par : ")
label_cheval2.place(x=450, y=330)

label_cheval3 = tk.Label(
    window, textvariable=varcheval)
label_cheval3.place(x=700, y=330)

cheval_listbox = tk.Listbox(window, height=len(plannig1.liste_chevaux))
cheval_listbox.place(x=200, y=50)


def items_selected_cheval(event):
    # get all selected indices
    cavalier = []
    selected_indices = cheval_listbox.curselection()
    cheval = cheval_listbox.get(selected_indices)
    cellule.set_cheval(cheval[1])
    ancient_cavalier = plannig1.ancient_cavalier_de(
        cheval_listbox.get(selected_indices)[1])
    for i in ancient_cavalier:
        cavalier.append(f"{i[0]} a {i[1]}")
    varheure_cheval.set(f"heure de travaille de : {cellule.cheval}")
    inserer_liste_de_travaille()
    varcheval.set(cavalier)
    varajout.set(cellule.getCellule())


cheval_listbox.bind('<ButtonRelease-1>', items_selected_cheval)

visu_fichier = tk.Text(window, width=100)
visu_fichier.place(x=700, y=360)

label_ajout = tk.Label(window, textvariable=varajout)
label_ajout.place(x=450, y=470)

boutton_ajouter = tk.Button(
    window, text="Ajouter", command=ajouter)
boutton_ajouter.place(x=450, y=500)

boutton_supprimer = tk.Button(
    window, text="supprimer", command=supprimer)
boutton_supprimer.place(x=520, y=500)

boutton_enregistrer = tk.Button(
    window, text="enregistrer", command=plannig1.ecrire_fichier, width=18)
boutton_enregistrer.place(x=450, y=530)

label_enregistrer = tk.Label(
    window, text="le fichier à bien été enregistré")
label_enregistrer.place(x=450, y=560)
label_enregistrer.config(fg="#f0f0f0")


label_heure_cheval = tk.Label(
    window, textvariable=varheure_cheval)
label_heure_cheval.place(x=800, y=40)

heure_listebox = tk.Listbox(window, width=25, height=5)
heure_listebox.place(x=800, y=70)


def items_selected_heure_cheval(event):
    selected_indices = heure_listebox.curselection()
    (h, p) = heure_listebox.get(selected_indices)
    if h != heure.h:
        heure.set_heure(h)
        cellule.set_ind_cava((None, None))
    else:
        Nb = 0
        for i in range(0, eleve_listbox.size()):
            if p == eleve_listbox.get(i):
                cellule.set_ind_cava((Nb, 0))
            Nb += 1
    cellule.set_cavalier(p)
    varajout.set(cellule.getCellule())


heure_listebox.bind('<ButtonRelease-1>', items_selected_heure_cheval)

label_historique = tk.Label(window, text="historique")
label_historique.place(x=1000, y=40)

historique = tk.Text(window, width=60, height=13)
historique.place(x=1000, y=70)

# créer un menu
menubar = Menu(window)
# créer un sous-menu
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="Mercredi", command=recup_donne_mercredi)
filemenu.add_command(label="Samedi", command=recup_donne_samedi)

# créer un sous-menu
filemenuheure = Menu(menubar, tearoff=0)
filemenuheurek = Menu(menubar, tearoff=0)


def ajout_des_commande_lena():
    filemenuheure.delete(0, "end")
    if journee.j == "Mercredi":
        filemenuheure.add_command(label="10H LENA", command=heure.h1)
        filemenuheure.add_command(label="13H30 LENA", command=heure.h2)
        filemenuheure.add_command(label="14H LENA", command=heure.h3)
        filemenuheure.add_command(label="15H LENA", command=heure.h4)
        filemenuheure.add_command(label="16H LENA", command=heure.h5)
        filemenuheure.add_command(label="17H LENA", command=heure.h6)
        filemenuheure.add_command(label="18H LENA", command=heure.h7)
        filemenuheure.add_command(label="19H LENA", command=heure.h8)
    else:
        filemenuheure.add_command(label="9h30 L", command=heure.h1)
        filemenuheure.add_command(label="10H30 L", command=heure.h3)
        filemenuheure.add_command(label="11H30 L", command=heure.h5)
        filemenuheure.add_command(label="12H00 L", command=heure.h7)
        filemenuheure.add_command(label="14H00 L", command=heure.h9)
        filemenuheure.add_command(label="15H00 L", command=heure.h11)
        filemenuheure.add_command(label="16H00 L", command=heure.h13)
        filemenuheure.add_command(label="17H00 L", command=heure.h15)

    filemenuheurek.delete(0, "end")
    if journee.j == "Mercredi":
        filemenuheurek.add_command(label="15H KARINE", command=heure.h9)
        filemenuheurek.add_command(label="17H KARINE", command=heure.h10)
        filemenuheurek.add_command(label="18H KARINE", command=heure.h11)
    else:
        filemenuheurek.add_command(label="9h30 K", command=heure.h2)
        filemenuheurek.add_command(label="10H30 K", command=heure.h4)
        filemenuheurek.add_command(label="11H30 K", command=heure.h6)
        filemenuheurek.add_command(label="13H00 K", command=heure.h8)
        filemenuheurek.add_command(label="14H00 K", command=heure.h10)
        filemenuheurek.add_command(label="16H00 K", command=heure.h12)
        filemenuheurek.add_command(label="17H00 K", command=heure.h14)
        filemenuheurek.add_command(label="18H00 K", command=heure.h16)


menubar.add_cascade(label="Jour", menu=filemenu)
menubar.add_cascade(label="Heure", menu=filemenuheure)
menubar.add_cascade(label="Heure karine", menu=filemenuheurek)
menubar.add_command(label="Quitter!", command=window.quit)

# afficher le menu
window.config(menu=menubar)

window.mainloop()
